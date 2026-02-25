from typing import Optional, List, Dict, Any
import io

from datetime import datetime, timedelta
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text
from starlette.concurrency import run_in_threadpool
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.core.deps import get_db, require_admin
from app.schemas.agents import ConversationExportRequest
from app.services.agents import analyze_hot_questions, analyze_student_chains

router = APIRouter()


@router.post("/admin/export/conversations")
async def export_selected_conversations_excel(
    payload: ConversationExportRequest,
    current_user: Dict[str, Any] = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    session_ids = [s for s in (payload.session_ids or []) if s]
    if not session_ids:
        raise HTTPException(status_code=400, detail="未提供 session_ids")
    if len(session_ids) > 200:
        raise HTTPException(status_code=400, detail="一次最多导出 200 个会话")

    sql = text(
        """
        SELECT
            v.id,
            v.session_id,
            v.user_id,
            v.display_user_name,
            u.student_id,
            u.class_name,
            u.study_year,
            v.agent_id,
            v.display_agent_name,
            v.message_type,
            v.content,
            v.response_time_ms,
            v.created_at
        FROM v_conversations_with_deleted v
        LEFT JOIN sys_users u ON v.user_id = u.id
        WHERE session_id = ANY(:session_ids)
        ORDER BY v.session_id ASC, v.created_at ASC, v.id ASC
        """
    )
    result = await db.execute(sql, {"session_ids": session_ids})
    rows = [dict(r) for r in result.mappings().all()]
    if not rows:
        raise HTTPException(status_code=404, detail="未找到对应会话的对话记录")

    def build_excel_bytes(items: List[Dict[str, Any]]) -> bytes:
        def fmt_time(value: Any) -> str:
            if value is None:
                return ""
            if hasattr(value, "isoformat"):
                s = value.isoformat(sep=" ", timespec="microseconds")
            else:
                s = str(value)
            return s.replace("T", " ").replace("Z", "").replace("+00:00", "")

        by_session: Dict[str, List[Dict[str, Any]]] = {}
        for r in items:
            sid = str(r.get("session_id") or "")
            if not sid:
                continue
            by_session.setdefault(sid, []).append(r)

        session_order = sorted(
            by_session.keys(),
            key=lambda sid: max(
                (row.get("created_at") for row in by_session[sid]), default=""
            ),
            reverse=True,
        )

        header = [
            "智能体名称",
            "学生姓名",
            "学号",
            "班级",
            "学年",
            "会话ID",
            "开始时间",
            "结束时间",
            "问题",
            "回答",
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "对话导出"

        header_fill = PatternFill("solid", fgColor="F2F2F2")
        header_font = Font(bold=True)
        header_alignment = Alignment(
            vertical="center", horizontal="center", wrap_text=True
        )
        left_alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        qa_alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)

        ws.append(header)
        for col_idx in range(1, len(header) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        ws.freeze_panes = "G2"

        col_widths = {
            "A": 18,
            "B": 14,
            "C": 14,
            "D": 14,
            "E": 10,
            "F": 38,
            "G": 22,
            "H": 22,
            "I": 60,
            "J": 60,
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        current_row = 2
        for sid in session_order:
            rows_for_session = sorted(
                by_session[sid], key=lambda r: (r.get("created_at"), r.get("id"))
            )
            agent_name = str(rows_for_session[0].get("display_agent_name") or "")
            user_name = str(rows_for_session[0].get("display_user_name") or "")
            student_id = str(rows_for_session[0].get("student_id") or "")
            class_name = str(rows_for_session[0].get("class_name") or "")
            study_year = str(rows_for_session[0].get("study_year") or "")

            pending_q: Optional[Dict[str, Any]] = None
            session_start_row = current_row

            for r in rows_for_session:
                mt = r.get("message_type")
                if mt == "question":
                    pending_q = r
                    continue
                if mt == "answer" and pending_q is not None:
                    start_at = fmt_time(pending_q.get("created_at"))
                    end_at = fmt_time(r.get("created_at"))
                    question = str(pending_q.get("content") or "")
                    answer = str(r.get("content") or "")
                    ws.append(
                        [
                            agent_name,
                            user_name,
                            student_id,
                            class_name,
                            study_year,
                            sid,
                            start_at,
                            end_at,
                            question,
                            answer,
                        ]
                    )
                    current_row += 1
                    pending_q = None

            if pending_q is not None:
                start_at = fmt_time(pending_q.get("created_at"))
                question = str(pending_q.get("content") or "")
                ws.append(
                    [
                        agent_name,
                        user_name,
                        student_id,
                        class_name,
                        study_year,
                        sid,
                        start_at,
                        "",
                        question,
                        "",
                    ]
                )
                current_row += 1

            session_end_row = current_row - 1
            if (
                session_end_row >= session_start_row
                and session_end_row > session_start_row
            ):
                for col_idx in range(1, 7):
                    ws.merge_cells(
                        start_row=session_start_row,
                        start_column=col_idx,
                        end_row=session_end_row,
                        end_column=col_idx,
                    )

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=10):
            for cell in row:
                if cell.column <= 8:
                    cell.alignment = left_alignment
                else:
                    cell.alignment = qa_alignment

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    excel_bytes = await run_in_threadpool(build_excel_bytes, rows)

    filename = f"conversations_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.get("/admin/export/hot-questions")
async def export_hot_questions_excel(
    agent_id: int = Query(..., ge=1, description="智能体ID"),
    start_at: Optional[datetime] = Query(None, description="开始时间(ISO)"),
    end_at: Optional[datetime] = Query(None, description="结束时间(ISO)"),
    bucket_seconds: int = Query(60, ge=30, le=900, description="时间桶(秒)，如60/180"),
    top_n: int = Query(10, ge=1, le=50, description="每个时间桶返回TopN问题"),
    current_user: Dict[str, Any] = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    now = datetime.utcnow()
    effective_end = end_at or now
    effective_start = start_at or (effective_end - timedelta(hours=1))
    rows = await analyze_hot_questions(
        db,
        agent_id=agent_id,
        start_at=effective_start,
        end_at=effective_end,
        bucket_seconds=bucket_seconds,
        top_n=top_n,
    )

    def build_excel_bytes(items: List[Dict[str, Any]]) -> bytes:
        def fmt_time(value: Any) -> str:
            if value is None:
                return ""
            if hasattr(value, "isoformat"):
                s = value.isoformat(sep=" ", timespec="microseconds")
            else:
                s = str(value)
            return s.replace("T", " ").replace("Z", "").replace("+00:00", "")

        header = ["时间桶开始", "问题总数", "参与学生数", "排名", "问题", "出现次数"]

        wb = Workbook()
        ws = wb.active
        ws.title = "热点问题"

        header_fill = PatternFill("solid", fgColor="F2F2F2")
        header_font = Font(bold=True)
        header_alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        left_alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
        center_alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

        ws.append(header)
        for col_idx in range(1, len(header) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        ws.freeze_panes = "A2"
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 8
        ws.column_dimensions["E"].width = 70
        ws.column_dimensions["F"].width = 12

        current_row = 2
        for bucket in sorted(items, key=lambda x: x.get("bucket_start") or ""):
            bucket_start = fmt_time(bucket.get("bucket_start"))
            question_count = int(bucket.get("question_count") or 0)
            unique_students = int(bucket.get("unique_students") or 0)
            top_questions = bucket.get("top_questions") or []

            start_row = current_row
            if not top_questions:
                ws.append([bucket_start, question_count, unique_students, "", "", ""])
                current_row += 1
            else:
                for idx, q in enumerate(top_questions, start=1):
                    ws.append(
                        [
                            bucket_start,
                            question_count,
                            unique_students,
                            idx,
                            str(q.get("question") or ""),
                            int(q.get("count") or 0),
                        ]
                    )
                    current_row += 1

            end_row = current_row - 1
            if end_row > start_row:
                for col_idx in range(1, 4):
                    ws.merge_cells(
                        start_row=start_row,
                        start_column=col_idx,
                        end_row=end_row,
                        end_column=col_idx,
                    )

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
            for cell in row:
                if cell.column in (1, 5):
                    cell.alignment = left_alignment
                else:
                    cell.alignment = center_alignment

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    excel_bytes = await run_in_threadpool(build_excel_bytes, rows)
    filename = f"hot_questions_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.get("/admin/export/student-chains")
async def export_student_chains_excel(
    agent_id: int = Query(..., ge=1, description="智能体ID"),
    user_id: Optional[int] = Query(None, ge=1, description="用户ID"),
    student_id: Optional[str] = Query(None, description="学号"),
    start_at: Optional[datetime] = Query(None, description="开始时间(ISO)"),
    end_at: Optional[datetime] = Query(None, description="结束时间(ISO)"),
    limit_sessions: int = Query(5, ge=1, le=20, description="最多返回会话数"),
    current_user: Dict[str, Any] = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    now = datetime.utcnow()
    effective_end = end_at or now
    effective_start = start_at or (effective_end - timedelta(hours=1))
    rows = await analyze_student_chains(
        db,
        agent_id=agent_id,
        user_id=user_id,
        student_id=student_id,
        start_at=effective_start,
        end_at=effective_end,
        limit_sessions=limit_sessions,
    )

    def build_excel_bytes(items: List[Dict[str, Any]]) -> bytes:
        def fmt_time(value: Any) -> str:
            if value is None:
                return ""
            if hasattr(value, "isoformat"):
                s = value.isoformat(sep=" ", timespec="microseconds")
            else:
                s = str(value)
            return s.replace("T", " ").replace("Z", "").replace("+00:00", "")

        header = ["会话ID", "最后时间", "轮次", "提问链条", "会话摘要"]

        wb = Workbook()
        ws = wb.active
        ws.title = "会话摘要"

        header_fill = PatternFill("solid", fgColor="F2F2F2")
        header_font = Font(bold=True)
        header_alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        left_alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
        center_alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

        ws.append(header)
        for col_idx in range(1, len(header) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        ws.freeze_panes = "A2"
        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 60
        ws.column_dimensions["E"].width = 90

        for s in items:
            sid = str(s.get("session_id") or "")
            last_at = fmt_time(s.get("last_at"))
            turns = int(s.get("turns") or 0)
            messages = s.get("messages") or []
            questions = [str(m.get("content") or "") for m in messages if m.get("message_type") == "question"]
            question_chain = "\n".join([q for q in questions if q.strip()])

            lines: List[str] = []
            for m in messages:
                mt = m.get("message_type")
                prefix = "问" if mt == "question" else "答" if mt == "answer" else str(mt or "")
                ts = fmt_time(m.get("created_at"))
                content = str(m.get("content") or "")
                if ts:
                    lines.append(f"{ts} {prefix}: {content}")
                else:
                    lines.append(f"{prefix}: {content}")
            summary_text = "\n".join([ln for ln in lines if ln.strip()])

            ws.append([sid, last_at, turns, question_chain, summary_text])

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
            for cell in row:
                if cell.column in (1, 2, 4, 5):
                    cell.alignment = left_alignment
                else:
                    cell.alignment = center_alignment

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    excel_bytes = await run_in_threadpool(build_excel_bytes, rows)
    filename = f"student_chains_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
