import io
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, select
from sqlalchemy.dialects.postgresql import insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.deps import require_admin
from app.db.database import get_db
from app.models import XbkCourse, XbkSelection, XbkStudent

router = APIRouter()


def _normalize_str(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _normalize_col_name(v: Any) -> str:
    return str(v or "").strip().replace("\u00a0", "").replace(" ", "")


def _parse_quota_int(value: Any) -> Optional[int]:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    try:
        return int(float(s))
    except Exception:
        return None


def _read_excel(file: UploadFile) -> pd.DataFrame:
    content = file.file.read()
    if not content:
        raise HTTPException(status_code=400, detail="文件为空")
    try:
        return pd.read_excel(io.BytesIO(content), dtype=str)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"读取Excel失败: {e}")


def _get_year_term(row: pd.Series, year: Optional[int], term: Optional[str]) -> Tuple[int, str]:
    y = row.get("年份")
    t = row.get("学期")
    final_year = year if year is not None else (_normalize_str(y) if pd.notna(y) else None)
    final_term = term if term else (_normalize_str(t) if pd.notna(t) else None)
    if final_year is None or not final_term:
        raise HTTPException(status_code=422, detail="缺少年份/学期（可通过查询参数传入或Excel列提供）")
    try:
        return int(str(final_year)), str(final_term)
    except Exception:
        raise HTTPException(status_code=422, detail="年份格式不正确")


def _remap_columns(df: pd.DataFrame, mapping: Dict[str, List[str]]) -> pd.DataFrame:
    cols = {_normalize_col_name(c): c for c in df.columns}
    rename: Dict[str, str] = {}
    for canonical, aliases in mapping.items():
        for name in [canonical, *aliases]:
            key = _normalize_col_name(name)
            if key in cols:
                rename[cols[key]] = canonical
                break
    return df.rename(columns=rename)


def _preview_rows(rows: List[Dict[str, Any]], limit: int = 10) -> List[Dict[str, Any]]:
    return rows[: max(0, min(limit, 50))]


def _style_worksheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="F0F2F5")
    header_font = Font(bold=True, color="000000")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_align = Alignment(vertical="center", wrap_text=True)

    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < 1 or max_col < 1:
        return

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    ws.row_dimensions[1].height = 22

    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for row in range(2, max_row + 1):
        ws.row_dimensions[row].height = 18
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).alignment = body_align

    for col in range(1, max_col + 1):
        max_len = 0
        for row in range(1, min(max_row, 200) + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        width = min(max(max_len + 2, 10), 48)
        ws.column_dimensions[get_column_letter(col)].width = width


def _students_mapping() -> Dict[str, List[str]]:
    return {
        "年份": ["year"],
        "学期": ["term"],
        "班级": ["班别", "班级名称", "class", "class_name"],
        "学号": ["学生学号", "student_no", "studentId", "student_id"],
        "姓名": ["学生姓名", "name", "student_name"],
        "性别": ["gender", "sex"],
    }


def _courses_mapping() -> Dict[str, List[str]]:
    return {
        "年份": ["year"],
        "学期": ["term"],
        "课程代码": ["代码", "course_code", "courseId", "course_id"],
        "课程名称": ["名称", "course_name"],
        "课程负责人": ["教师", "任课老师", "teacher"],
        "各班限报人数": ["限报人数", "quota_by_class", "quota"],
        "上课地点": ["地点", "location", "classroom"],
    }


def _selections_mapping() -> Dict[str, List[str]]:
    return {
        "年份": ["year"],
        "学期": ["term"],
        "学号": ["学生学号", "student_no", "studentId", "student_id"],
        "姓名": ["学生姓名", "name", "student_name"],
        "课程代码": ["代码", "course_code", "courseId", "course_id"],
    }


def _validate_required_columns(df: pd.DataFrame, required: List[str]) -> None:
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise HTTPException(status_code=422, detail=f"缺少必要列: {', '.join(missing)}")


def _row_errors(idx: int, messages: List[str]) -> Dict[str, Any]:
    return {"row": idx, "errors": messages}


@router.get("/import/template")
async def download_template(
    scope: str = Query(..., pattern="^(students|courses|selections)$"),
    db: AsyncSession = Depends(get_db),
    _: Dict[str, Any] = Depends(require_admin),
) -> StreamingResponse:
    if scope == "students":
        df = pd.DataFrame(
            [
                {"年份": 2026, "学期": "上学期", "班级": "高一(1)班", "学号": "20260001", "姓名": "张三", "性别": "男"},
            ]
        )
    elif scope == "courses":
        df = pd.DataFrame(
            [
                {
                    "年份": 2026,
                    "学期": "上学期",
                    "课程代码": "12",
                    "课程名称": "Python 基础与应用",
                    "课程负责人": "王老师",
                    "各班限报人数": 3,
                    "上课地点": "艺术中心1楼舞蹈教室2",
                },
            ]
        )
    else:
        df = pd.DataFrame(
            [
                {"年份": 2026, "学期": "上学期", "学号": "20260001", "姓名": "张三", "课程代码": "12"},
            ]
        )

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="template")
    output.seek(0)
    headers = {"Content-Disposition": f'attachment; filename="xbk_{scope}_template.xlsx"'}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.post("/import/preview")
async def preview_import(
    scope: str = Query(..., pattern="^(students|courses|selections)$"),
    year: Optional[int] = Query(None),
    term: Optional[str] = Query(None),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _: Dict[str, Any] = Depends(require_admin),
) -> Dict[str, Any]:
    df = _read_excel(file)
    df = df.fillna("")

    mapping = _students_mapping() if scope == "students" else _courses_mapping() if scope == "courses" else _selections_mapping()
    df = _remap_columns(df, mapping)

    if scope == "students":
        required = ["班级", "学号", "姓名"]
    elif scope == "courses":
        required = ["课程代码", "课程名称"]
    else:
        required = ["学号", "课程代码"]
    _validate_required_columns(df, required)

    total_rows = int(df.shape[0])
    errors: List[Dict[str, Any]] = []
    valid_rows: List[Dict[str, Any]] = []
    invalid = 0

    for i, (_, row) in enumerate(df.iterrows(), start=2):
        row_dict = {k: _normalize_str(row.get(k)) for k in df.columns}
        messages: List[str] = []
        try:
            _get_year_term(row, year, term)
        except HTTPException as e:
            messages.append(str(e.detail))
        if scope == "students":
            if not row_dict.get("班级"):
                messages.append("班级不能为空")
            if not row_dict.get("学号"):
                messages.append("学号不能为空")
            if not row_dict.get("姓名"):
                messages.append("姓名不能为空")
        elif scope == "courses":
            if not row_dict.get("课程代码"):
                messages.append("课程代码不能为空")
            if not row_dict.get("课程名称"):
                messages.append("课程名称不能为空")
            quota = _parse_quota_int(row_dict.get("各班限报人数"))
            if row_dict.get("各班限报人数") and quota is None:
                messages.append("限报人数格式不正确（应为数字）")
        else:
            if not row_dict.get("学号"):
                messages.append("学号不能为空")
            if not row_dict.get("课程代码"):
                messages.append("课程代码不能为空")

        if messages:
            invalid += 1
            errors.append(_row_errors(i, messages))
        else:
            valid_rows.append(row_dict)

    return {
        "total_rows": total_rows,
        "valid_rows": total_rows - invalid,
        "invalid_rows": invalid,
        "errors": errors[:50],
        "preview": _preview_rows(valid_rows, 10),
        "columns": list(df.columns),
    }


@router.post("/import", status_code=200)
async def import_data(
    scope: str = Query(..., pattern="^(students|courses|selections)$"),
    year: Optional[int] = Query(None),
    term: Optional[str] = Query(None),
    grade: Optional[str] = Query(None),
    skip_invalid: bool = Query(True),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _: Dict[str, Any] = Depends(require_admin),
) -> Dict[str, Any]:
    df = _read_excel(file)
    if df.empty:
        return {"total_rows": 0, "processed": 0, "inserted": 0, "updated": 0, "skipped": 0, "invalid": 0, "errors": []}

    df = df.fillna("")
    now = datetime.utcnow()
    mapping = _students_mapping() if scope == "students" else _courses_mapping() if scope == "courses" else _selections_mapping()
    df = _remap_columns(df, mapping)

    # 优先读取 Excel 中的年级列，如果没有则使用 Query 参数
    def _get_row_grade(row, default_grade):
        g = _normalize_str(row.get("年级"))
        return g if g else default_grade

    if scope == "students":
        required = ["班级", "学号", "姓名"]
        _validate_required_columns(df, required)
        keys = []
        for _, row in df.iterrows():
            try:
                y, t = _get_year_term(row, year, term)
            except HTTPException:
                continue
            student_no = _normalize_str(row.get("学号"))
            if student_no:
                keys.append((y, t, student_no))
        existing: set[Tuple[int, str, str]] = set()
        if keys:
            year_terms = {(k[0], k[1]) for k in keys}
            for y, t in year_terms:
                stmt = select(XbkStudent.year, XbkStudent.term, XbkStudent.student_no).where(
                    XbkStudent.is_deleted.is_(False), XbkStudent.year == y, XbkStudent.term == t
                )
                for row in (await db.execute(stmt)).all():
                    existing.add((int(row[0]), str(row[1]), str(row[2])))

        processed = inserted = updated = skipped = invalid = 0
        errors: List[Dict[str, Any]] = []

        for i, (_, row) in enumerate(df.iterrows(), start=2):
            row_errors: List[str] = []
            try:
                y, t = _get_year_term(row, year, term)
            except HTTPException as e:
                row_errors.append(str(e.detail))
                y = t = None
            
            row_grade = _get_row_grade(row, grade)
            class_name = _normalize_str(row.get("班级"))
            student_no = _normalize_str(row.get("学号"))
            name = _normalize_str(row.get("姓名"))
            gender = _normalize_str(row.get("性别"))
            if not class_name:
                row_errors.append("班级不能为空")
            if not student_no:
                row_errors.append("学号不能为空")
            if not name:
                row_errors.append("姓名不能为空")

            if row_errors:
                invalid += 1
                if not skip_invalid:
                    raise HTTPException(status_code=422, detail={"row": i, "errors": row_errors})
                errors.append(_row_errors(i, row_errors))
                continue

            stmt = (
                insert(XbkStudent)
                .values(
                    year=y,
                    term=t,
                    grade=row_grade,
                    class_name=class_name,
                    student_no=student_no,
                    name=name,
                    gender=gender,
                    is_deleted=False,
                    created_at=now,
                    updated_at=now,
                )
                .on_conflict_do_update(
                    index_elements=["year", "term", "student_no"],
                    set_={
                        "grade": row_grade,
                        "class_name": class_name,
                        "name": name,
                        "gender": gender,
                        "is_deleted": False,
                        "updated_at": now,
                    },
                )
            )
            await db.execute(stmt)
            processed += 1
            if (y, t, student_no) in existing:
                updated += 1
            else:
                inserted += 1
        await db.commit()
        return {
            "total_rows": int(df.shape[0]),
            "processed": processed,
            "inserted": inserted,
            "updated": updated,
            "skipped": skipped,
            "invalid": invalid,
            "errors": errors[:50],
        }

    if scope == "courses":
        required = ["课程代码", "课程名称"]
        _validate_required_columns(df, required)
        keys = []
        for _, row in df.iterrows():
            try:
                y, t = _get_year_term(row, year, term)
            except HTTPException:
                continue
            course_code = _normalize_str(row.get("课程代码"))
            if course_code:
                keys.append((y, t, course_code))
        existing: set[Tuple[int, str, str]] = set()
        if keys:
            year_terms = {(k[0], k[1]) for k in keys}
            for y, t in year_terms:
                stmt = select(XbkCourse.year, XbkCourse.term, XbkCourse.course_code).where(
                    XbkCourse.is_deleted.is_(False), XbkCourse.year == y, XbkCourse.term == t
                )
                for row in (await db.execute(stmt)).all():
                    existing.add((int(row[0]), str(row[1]), str(row[2])))

        processed = inserted = updated = skipped = invalid = 0
        errors: List[Dict[str, Any]] = []

        for i, (_, row) in enumerate(df.iterrows(), start=2):
            row_errors: List[str] = []
            try:
                y, t = _get_year_term(row, year, term)
            except HTTPException as e:
                row_errors.append(str(e.detail))
                y = t = None
            
            row_grade = _get_row_grade(row, grade)
            course_code = _normalize_str(row.get("课程代码"))
            course_name = _normalize_str(row.get("课程名称"))
            teacher = _normalize_str(row.get("课程负责人"))
            quota_value = _normalize_str(row.get("各班限报人数"))
            location = _normalize_str(row.get("上课地点"))
            if not course_code:
                row_errors.append("课程代码不能为空")
            if not course_name:
                row_errors.append("课程名称不能为空")
            quota = _parse_quota_int(quota_value)
            if quota_value and quota is None:
                row_errors.append("限报人数格式不正确（应为数字）")
                quota = 0

            if row_errors:
                invalid += 1
                if not skip_invalid:
                    raise HTTPException(status_code=422, detail={"row": i, "errors": row_errors})
                errors.append(_row_errors(i, row_errors))
                continue

            stmt = (
                insert(XbkCourse)
                .values(
                    year=y,
                    term=t,
                    grade=row_grade,
                    course_code=course_code,
                    course_name=course_name,
                    teacher=teacher,
                    quota=quota or 0,
                    location=location,
                    is_deleted=False,
                    created_at=now,
                    updated_at=now,
                )
                .on_conflict_do_update(
                    index_elements=["year", "term", "course_code"],
                    set_={
                        "grade": row_grade,
                        "course_name": course_name,
                        "teacher": teacher,
                        "quota": quota or 0,
                        "location": location,
                        "is_deleted": False,
                        "updated_at": now,
                    },
                )
            )
            await db.execute(stmt)
            processed += 1
            if (y, t, course_code) in existing:
                updated += 1
            else:
                inserted += 1
        await db.commit()
        return {
            "total_rows": int(df.shape[0]),
            "processed": processed,
            "inserted": inserted,
            "updated": updated,
            "skipped": skipped,
            "invalid": invalid,
            "errors": errors[:50],
        }

    required = ["学号", "课程代码"]
    _validate_required_columns(df, required)
    keys = []
    for _, row in df.iterrows():
        try:
            y, t = _get_year_term(row, year, term)
        except HTTPException:
            continue
        student_no = _normalize_str(row.get("学号"))
        course_code = _normalize_str(row.get("课程代码"))
        if student_no and course_code:
            keys.append((y, t, student_no, course_code))
    existing: set[Tuple[int, str, str, str]] = set()
    if keys:
        year_terms = {(k[0], k[1]) for k in keys}
        for y, t in year_terms:
            stmt = select(XbkSelection.year, XbkSelection.term, XbkSelection.student_no, XbkSelection.course_code).where(
                XbkSelection.is_deleted.is_(False), XbkSelection.year == y, XbkSelection.term == t
            )
            for row in (await db.execute(stmt)).all():
                existing.add((int(row[0]), str(row[1]), str(row[2]), str(row[3])))

    processed = inserted = updated = skipped = invalid = 0
    errors: List[Dict[str, Any]] = []

    for i, (_, row) in enumerate(df.iterrows(), start=2):
        row_errors: List[str] = []
        try:
            y, t = _get_year_term(row, year, term)
        except HTTPException as e:
            row_errors.append(str(e.detail))
            y = t = None
        
        row_grade = _get_row_grade(row, grade)
        student_no = _normalize_str(row.get("学号"))
        name = _normalize_str(row.get("姓名"))
        course_code = _normalize_str(row.get("课程代码"))
        if not student_no:
            row_errors.append("学号不能为空")
        if not course_code:
            row_errors.append("课程代码不能为空")
        if row_errors:
            invalid += 1
            if not skip_invalid:
                raise HTTPException(status_code=422, detail={"row": i, "errors": row_errors})
            errors.append(_row_errors(i, row_errors))
            continue

        stmt = (
            insert(XbkSelection)
            .values(
                year=y,
                term=t,
                grade=row_grade,
                student_no=student_no,
                name=name,
                course_code=course_code,
                is_deleted=False,
                created_at=now,
                updated_at=now,
            )
            .on_conflict_do_update(
                index_elements=["year", "term", "student_no", "course_code"],
                set_={"grade": row_grade, "name": name, "is_deleted": False, "updated_at": now},
            )
        )
        await db.execute(stmt)
        processed += 1
        if (y, t, student_no, course_code) in existing:
            updated += 1
        else:
            inserted += 1
    await db.commit()
    return {
        "total_rows": int(df.shape[0]),
        "processed": processed,
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "invalid": invalid,
        "errors": errors[:50],
    }


@router.get("/export")
async def export_data(
    scope: str = Query(..., pattern="^(students|courses|selections)$"),
    year: Optional[int] = Query(None),
    term: Optional[str] = Query(None),
    class_name: Optional[str] = Query(None),
    format: str = Query("xlsx", pattern="^(xlsx|xls)$"),
    db: AsyncSession = Depends(get_db),
    _: Dict[str, Any] = Depends(require_admin),
) -> StreamingResponse:
    if scope == "students":
        stmt = select(XbkStudent).where(XbkStudent.is_deleted.is_(False))
        if year is not None:
            stmt = stmt.where(XbkStudent.year == year)
        if term:
            stmt = stmt.where(XbkStudent.term == term)
        if class_name:
            stmt = stmt.where(XbkStudent.class_name == class_name)
        rows = (await db.execute(stmt.order_by(XbkStudent.class_name.asc(), XbkStudent.student_no.asc()))).scalars().all()
        df = pd.DataFrame(
            [
                {
                    "年份": r.year,
                    "学期": r.term,
                    "班级": r.class_name,
                    "学号": r.student_no,
                    "姓名": r.name,
                    "性别": r.gender,
                }
                for r in rows
            ]
        )
    elif scope == "courses":
        stmt = select(XbkCourse).where(XbkCourse.is_deleted.is_(False))
        if year is not None:
            stmt = stmt.where(XbkCourse.year == year)
        if term:
            stmt = stmt.where(XbkCourse.term == term)
        rows = (await db.execute(stmt.order_by(XbkCourse.course_code.asc()))).scalars().all()
        df = pd.DataFrame(
            [
                {
                    "年份": r.year,
                    "学期": r.term,
                    "课程代码": r.course_code,
                    "课程名称": r.course_name,
                    "课程负责人": r.teacher,
                    "各班限报人数": r.quota,
                    "上课地点": r.location,
                }
                for r in rows
            ]
        )
    else:
        stmt = select(XbkSelection).where(XbkSelection.is_deleted.is_(False))
        if year is not None:
            stmt = stmt.where(XbkSelection.year == year)
        if term:
            stmt = stmt.where(XbkSelection.term == term)
        if class_name:
            sub = select(XbkStudent.student_no).where(
                XbkStudent.is_deleted.is_(False),
                XbkStudent.class_name == class_name,
                *( [XbkStudent.year == year] if year is not None else [] ),
                *( [XbkStudent.term == term] if term else [] ),
            )
            stmt = stmt.where(XbkSelection.student_no.in_(sub))
        rows = (await db.execute(stmt.order_by(XbkSelection.student_no.asc(), XbkSelection.course_code.asc()))).scalars().all()
        df = pd.DataFrame(
            [
                {
                    "年份": r.year,
                    "学期": r.term,
                    "学号": r.student_no,
                    "姓名": r.name,
                    "课程代码": r.course_code,
                }
                for r in rows
            ]
        )

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="data")
        ws = writer.book["data"]
        _style_worksheet(ws)
    output.seek(0)

    filename = f"xbk_{scope}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{format}"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
