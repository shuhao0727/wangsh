"""
用户管理 API 端点
与 sys_users 表交互，提供用户数据的 CRUD 操作
支持管理员管理所有用户，包括学生和管理员
"""

from typing import Any, Dict, Iterator, List, Optional, Tuple, cast
from datetime import datetime
import csv
import io
from urllib.parse import quote
from fastapi import APIRouter, Depends, HTTPException, status, Query, Body, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import select, or_, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.exc import IntegrityError
from pydantic import BaseModel, Field, ConfigDict

from app.core.deps import require_admin, get_db
from app.models import User
from app.core.pubsub import publish

router = APIRouter()

USER_IMPORT_HEADERS = ["学号", "姓名", "学年", "班级", "状态", "用户名"]
USER_IMPORT_REQUIRED_FIELDS = ["学号", "姓名"]
USER_IMPORT_TEMPLATE_ROWS = [
    ["20230001", "张三", "2025", "高一(1)班", "true", "zhangsan"],
    ["20230002", "李四", "2025", "高一(2)班", "true", "lisi"],
    ["20230003", "王五", "2025", "高一(3)班", "false", "wangwu"],
    ["20230004", "赵六", "2025", "高一(1)班", "true", ""],
]


def _normalize_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value).strip()


def _parse_csv_rows(content: bytes) -> Tuple[List[str], Iterator[Dict[str, str]]]:
    content_text = content.decode("utf-8-sig")
    csv_file = io.StringIO(content_text)
    reader = csv.DictReader(csv_file)
    if not reader.fieldnames:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="CSV文件为空或格式不正确",
        )

    fieldnames = [str(name).strip() for name in reader.fieldnames if name]
    return fieldnames, reader


def _parse_xlsx_rows(content: bytes) -> Tuple[List[str], Iterator[Dict[str, str]]]:
    workbook = load_workbook(filename=io.BytesIO(content), data_only=True, read_only=True)
    worksheet = workbook.active
    assert worksheet is not None  # XLSX 文件总是有 active sheet
    row_iterator = worksheet.iter_rows(values_only=True)
    header_row = next(row_iterator, None)
    if not header_row:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel文件为空或格式不正确",
        )

    fieldnames = [_normalize_cell_value(value) for value in header_row]
    if not any(fieldnames):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel文件缺少表头",
        )

    def row_generator() -> Iterator[Dict[str, str]]:
        for row in row_iterator:
            current_row = row or ()
            parsed_row: Dict[str, str] = {}
            for idx, header in enumerate(fieldnames):
                if not header:
                    continue
                value = current_row[idx] if idx < len(current_row) else ""
                parsed_row[header] = _normalize_cell_value(value)
            yield parsed_row

    return [header for header in fieldnames if header], row_generator()


def _parse_import_rows(file_name: str, content: bytes) -> Tuple[List[str], Iterator[Dict[str, str]]]:
    lower_name = file_name.lower()
    if lower_name.endswith((".csv", ".txt")):
        return _parse_csv_rows(content)
    if lower_name.endswith(".xlsx"):
        return _parse_xlsx_rows(content)
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="只支持 CSV、TXT 或 XLSX 文件",
    )


# Pydantic 模型定义
class UserCreate(BaseModel):
    """用户创建请求模型"""
    student_id: Optional[str] = Field(None, max_length=50, description="学号")
    username: Optional[str] = Field(None, min_length=3, max_length=50, description="用户名")
    full_name: str = Field(..., max_length=100, description="全名")
    class_name: Optional[str] = Field(None, max_length=50, description="班级名称")
    study_year: Optional[str] = Field(None, max_length=10, description="学年")
    role_code: Optional[str] = Field("student", max_length=20, description="角色代码")
    is_active: Optional[bool] = Field(True, description="是否激活")


class UserUpdate(BaseModel):
    """用户更新请求模型"""
    student_id: Optional[str] = Field(None, max_length=50, description="学号")
    username: Optional[str] = Field(None, min_length=3, max_length=50, description="用户名")
    full_name: Optional[str] = Field(None, max_length=100, description="全名")
    class_name: Optional[str] = Field(None, max_length=50, description="班级名称")
    study_year: Optional[str] = Field(None, max_length=10, description="学年")
    role_code: Optional[str] = Field(None, max_length=20, description="角色代码")
    is_active: Optional[bool] = Field(None, description="是否激活")


class UserResponse(BaseModel):
    """用户响应模型"""
    id: int
    student_id: Optional[str]
    username: Optional[str]
    full_name: str
    class_name: Optional[str]
    study_year: Optional[str]
    role_code: str
    is_active: bool
    created_at: Optional[datetime]
    updated_at: Optional[datetime]
    
    model_config = ConfigDict(from_attributes=True)


class UserListResponse(BaseModel):
    """用户列表响应模型"""
    users: List[UserResponse]
    total: int
    skip: int
    limit: int
    has_more: bool


@router.get("/", response_model=UserListResponse)
async def list_users(
    skip: int = Query(0, ge=0, description="跳过记录数"),
    limit: int = Query(20, ge=1, le=100, description="每页记录数"),
    search: Optional[str] = Query(None, description="搜索关键词（学号、姓名、班级）"),
    role_code: Optional[str] = Query(None, description="角色代码过滤"),
    is_active: Optional[bool] = Query(None, description="是否激活状态过滤"),
    current_user = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
) -> UserListResponse:
    """
    获取用户列表（需要管理员权限）
    支持分页、搜索和过滤
    默认排除管理员角色（admin 和 super_admin）
    """
    try:
        # 构建查询条件
        conditions = []
        # 使用SQLAlchemy正确的语法
        conditions.append(User.is_deleted == False)
        
        # 默认排除管理员角色（admin 和 super_admin）
        # 只有当用户显式指定 role_code 时才不过滤
        if not role_code:
            conditions.append(User.role_code.notin_(['admin', 'super_admin']))
        
        if search:
            search_term = f"%{search}%"
            conditions.append(
                or_(
                    User.student_id.ilike(search_term),
                    User.full_name.ilike(search_term),
                    User.class_name.ilike(search_term),
                    User.username.ilike(search_term)
                )
            )
        
        if role_code:
            conditions.append(User.role_code == role_code)
        
        if is_active is not None:
            conditions.append(User.is_active == is_active)
        
        # 获取总数
        count_query = select(func.count()).select_from(User)
        if conditions:
            count_query = count_query.where(*conditions)
        count_result = await db.execute(count_query)
        total = count_result.scalar_one()
        
        # 获取分页数据
        query = select(User)
        if conditions:
            query = query.where(*conditions)
        query = query.order_by(User.created_at.desc()).offset(skip).limit(limit)
        
        result = await db.execute(query)
        users = result.scalars().all()
        
        # 转换为响应格式
        user_list = []
        for user in users:
            # 使用 Pydantic 的 model_validate 方法，正确处理 SQLAlchemy 对象
            user_response = UserResponse.model_validate(user)
            user_list.append(user_response)
        
        return UserListResponse(
            users=user_list,
            total=total,
            skip=skip,
            limit=limit,
            has_more=skip + limit < total
        )
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取用户列表失败: {str(e)}"
        )


@router.get("/{user_id}", response_model=UserResponse)
async def get_user(
    user_id: int,
    current_user = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
) -> UserResponse:
    """
    获取用户详情（需要管理员权限）
    """
    try:
        # 使用SQLAlchemy正确的语法
        query = select(User).where(
            User.id == user_id,
            User.is_deleted == False
        )
        result = await db.execute(query)
        user = result.scalar_one_or_none()
        
        if not user:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="用户不存在"
            )
        
        # 使用 Pydantic 的 model_validate 方法
        return UserResponse.model_validate(user)
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取用户详情失败: {str(e)}"
        )


@router.post("/", response_model=UserResponse)
async def create_user(
    user_data: UserCreate,
    current_user = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
) -> UserResponse:
    """
    创建新用户（需要管理员权限）
    """
    try:
        # 检查唯一性约束 - 检查值是否为None，而不是SQLAlchemy对象
        existing_checks = []
        
        if user_data.username is not None:
            existing_checks.append(User.username == user_data.username)
        
        if user_data.student_id is not None:
            existing_checks.append(User.student_id == user_data.student_id)
        
        # 类型忽略：Pylance不理解这个条件检查
        if existing_checks:  # type: ignore
            check_query = select(User).where(or_(*existing_checks))
            check_result = await db.execute(check_query)
            existing_user = check_result.scalar_one_or_none()
            
            if existing_user:
                if existing_user.username == user_data.username:  # type: ignore[union-attr]
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="用户名已存在"
                    )
                if existing_user.student_id == user_data.student_id:  # type: ignore[union-attr]
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="学号已存在"
                    )
        
        # 创建新用户
        new_user = User(
            student_id=user_data.student_id,
            username=user_data.username,
            full_name=user_data.full_name,
            class_name=user_data.class_name,
            study_year=user_data.study_year,
            role_code=user_data.role_code or "student",
            is_active=user_data.is_active if user_data.is_active is not None else True
        )
        
        db.add(new_user)
        await db.commit()
        await db.refresh(new_user)

        # 发布事件
        publish("admin_global", {"type": "user_changed", "action": "create", "id": new_user.id})

        # 使用 Pydantic 的 model_validate 方法
        return UserResponse.model_validate(new_user)
        
    except HTTPException:
        raise
    except IntegrityError as e:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="用户数据不符合数据库约束"
        )
    except Exception as e:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"创建用户失败: {str(e)}"
        )


@router.put("/{user_id}", response_model=UserResponse)
async def update_user(
    user_id: int,
    user_data: UserUpdate,
    current_user = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
) -> UserResponse:
    """
    更新用户信息（需要管理员权限）
    """
    try:
        # 获取现有用户 - 使用SQLAlchemy正确的语法
        query = select(User).where(
            User.id == user_id,
            User.is_deleted == False
        )
        result = await db.execute(query)
        user = result.scalar_one_or_none()
        
        if not user:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="用户不存在"
            )
        
        # 检查唯一性约束（排除当前用户）
        existing_checks = []
        
        # 检查值是否为None，而不是SQLAlchemy对象
        # 类型忽略：Pylance不理解这是Python值而不是SQLAlchemy对象
        if user_data.username is not None and user_data.username != user.username:
            existing_checks.append(User.username == user_data.username)
        
        if user_data.student_id is not None and user_data.student_id != user.student_id:
            existing_checks.append(User.student_id == user_data.student_id)
        
        # 类型忽略：Pylance不理解这个条件检查
        if existing_checks:  # type: ignore
            check_query = select(User).where(
                or_(*existing_checks),
                User.id != user_id
            )
            check_result = await db.execute(check_query)
            existing_user = check_result.scalar_one_or_none()
            
            if existing_user:
                if existing_user.username == user_data.username:  # type: ignore[union-attr]
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="用户名已存在"
                    )
                if existing_user.student_id == user_data.student_id:  # type: ignore[union-attr]
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="学号已存在"
                    )
        
        # 更新用户信息
        # 类型忽略：Pylance不理解SQLAlchemy的动态类型转换
        if user_data.student_id is not None:
            user.student_id = user_data.student_id  # type: ignore
        
        if user_data.username is not None:
            user.username = user_data.username  # type: ignore
        
        if user_data.full_name is not None:
            user.full_name = user_data.full_name  # type: ignore
        
        if user_data.class_name is not None:
            user.class_name = user_data.class_name  # type: ignore
        
        if user_data.study_year is not None:
            user.study_year = user_data.study_year  # type: ignore
        
        if user_data.role_code is not None:
            user.role_code = user_data.role_code  # type: ignore
        
        if user_data.is_active is not None:
            user.is_active = user_data.is_active  # type: ignore

        await db.commit()
        await db.refresh(user)

        # 发布事件
        publish("admin_global", {"type": "user_changed", "action": "update", "id": user_id})

        # 使用 Pydantic 的 model_validate 方法
        return UserResponse.model_validate(user)
        
    except HTTPException:
        raise
    except IntegrityError as e:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="用户数据不符合数据库约束"
        )
    except Exception as e:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"更新用户失败: {str(e)}"
        )


@router.delete("/{user_id}")
async def delete_user(
    user_id: int,
    current_user = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
) -> dict:
    """
    删除用户（软删除，需要管理员权限）
    """
    try:
        # 获取现有用户 - 使用SQLAlchemy正确的语法
        query = select(User).where(
            User.id == user_id,
            User.is_deleted == False
        )
        result = await db.execute(query)
        user = result.scalar_one_or_none()
        
        if not user:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="用户不存在"
            )
        
        # 软删除：标记为已删除
        # 类型忽略：Pylance不理解SQLAlchemy的动态类型转换
        user.is_deleted = True  # type: ignore
        await db.commit()

        # 发布事件
        publish("admin_global", {"type": "user_changed", "action": "delete", "id": user_id})

        return {
            "success": True,
            "message": "用户删除成功",
            "user_id": user_id
        }
        
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"删除用户失败: {str(e)}"
        )


@router.post("/batch-delete")
async def batch_delete_users(
    user_ids: List[int] = Body(...),
    current_user = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
) -> dict:
    """
    批量删除用户（软删除，需要管理员权限）
    """
    try:
        # 获取符合条件的用户 - 使用SQLAlchemy正确的语法
        query = select(User).where(
            User.id.in_(user_ids),
            User.is_deleted == False
        )
        result = await db.execute(query)
        users = result.scalars().all()
        
        if not users:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="未找到要删除的用户"
            )
        
        # 批量软删除
        deleted_ids = []
        for user in users:
            # 类型忽略：Pylance不理解SQLAlchemy的动态类型转换
            user.is_deleted = True  # type: ignore
            deleted_ids.append(user.id)
        
        await db.commit()

        publish("admin_global", {"type": "user_changed", "action": "batch_delete"})

        return {
            "success": True,
            "message": f"成功删除 {len(deleted_ids)} 个用户",
            "deleted_ids": deleted_ids
        }
        
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"批量删除用户失败: {str(e)}"
        )


class ImportUserResponse(BaseModel):
    """单行导入响应模型"""
    row_number: int
    student_id: Optional[str] = None
    full_name: str
    status: str  # success, error
    message: Optional[str] = None
    user_id: Optional[int] = None


class UserImportResult(BaseModel):
    """用户导入结果模型"""
    success: bool
    message: str
    total_rows: int
    imported_count: int = 0
    updated_count: int = 0
    error_count: int = 0
    errors: List[ImportUserResponse] = []


@router.get("/import/template")
async def download_user_import_template(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    current_user = Depends(require_admin),
) -> StreamingResponse:
    """
    下载用户导入模板，支持 xlsx / csv。
    """
    file_name = f"user_import_template.{format}"
    encoded_name = quote(file_name)
    headers = {"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_name}"}

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(USER_IMPORT_HEADERS)
        writer.writerow(["# 注意：只能导入学生用户，不允许导入管理员"])
        writer.writerow(["# 状态：true=激活，false=未激活（可选，默认为true）"])
        writer.writerow(["# 用户名：可选字段，如果不填将使用学号作为登录账号"])
        writer.writerow(["# 示例数据："])
        for sample_row in USER_IMPORT_TEMPLATE_ROWS:
            writer.writerow(sample_row)

        csv_bytes = output.getvalue().encode("utf-8-sig")
        return StreamingResponse(
            io.BytesIO(csv_bytes),
            media_type="text/csv; charset=utf-8",
            headers=headers,
        )

    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None  # Workbook() 总是有一个 active sheet
    worksheet.title = "用户导入模板"
    worksheet.append(USER_IMPORT_HEADERS)
    for sample_row in USER_IMPORT_TEMPLATE_ROWS:
        worksheet.append(sample_row)

    notes_sheet = workbook.create_sheet(title="填写说明")
    notes_sheet.append(["说明"])
    notes_sheet.append(["1. 只能导入学生用户，不允许导入管理员"])
    notes_sheet.append(["2. 状态字段可填 true/false（默认为 true）"])
    notes_sheet.append(["3. 用户名字段可选，不填将使用学号作为登录账号"])
    notes_sheet.append(["4. 学号、姓名为必填字段"])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.post("/import", response_model=UserImportResult)
async def import_users(
    file: UploadFile,
    current_user = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
) -> UserImportResult:
    """
    批量导入用户（CSV / XLSX 格式，需要管理员权限）
    注意：只能导入学生用户，不允许导入管理员
    """
    try:
        # 检查文件类型
        if not file.filename:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="文件名不能为空"
            )

        # 读取文件内容
        content = await file.read()
        fieldnames, reader = _parse_import_rows(file.filename, content)

        # 验证必要字段
        for field in USER_IMPORT_REQUIRED_FIELDS:
            if field not in fieldnames:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"导入文件缺少必要字段: {field}"
                )
        
        results = []
        imported_count = 0
        updated_count = 0
        error_count = 0
        
        # 按行处理
        for i, row in enumerate(reader, start=1):
            try:
                # 跳过注释行
                if any(str(value).startswith('#') for value in row.values() if value):
                    continue
                
                # 提取数据
                student_id = row.get('学号', '').strip()
                full_name = row.get('姓名', '').strip()
                study_year = row.get('学年', '').strip() or None
                class_name = row.get('班级', '').strip() or None
                status_str = row.get('状态', 'true').strip().lower()
                username = row.get('用户名', '').strip() or None
                
                # 验证必填字段
                if not student_id or not full_name:
                    raise ValueError("学号和姓名为必填字段")
                
                # 验证状态
                is_active = True
                if status_str in ['false', '0', 'no', '否']:
                    is_active = False
                elif status_str not in ['true', '1', 'yes', '是', '']:
                    raise ValueError(f"状态值无效: {status_str}")
                
                # 验证角色代码（不允许导入管理员）
                role_code = "student"
                
                # 检查是否已存在（按学号）
                existing_query = select(User).where(
                    User.student_id == student_id,
                    User.is_deleted == False
                )
                result = await db.execute(existing_query)
                existing_user = result.scalar_one_or_none()
                
                if existing_user:
                    # 更新现有用户
                    existing_user.full_name = full_name  # type: ignore[assignment]
                    existing_user.study_year = study_year or existing_user.study_year  # type: ignore[assignment]
                    existing_user.class_name = class_name or existing_user.class_name  # type: ignore[assignment]
                    existing_user.is_active = is_active  # type: ignore[assignment]
                    if username and username != existing_user.username:
                        # 检查用户名是否已被其他用户使用
                        username_check = select(User).where(
                            User.username == username,
                            User.id != existing_user.id,
                            User.is_deleted == False
                        )
                        username_result = await db.execute(username_check)
                        if username_result.scalar_one_or_none():
                            raise ValueError(f"用户名 '{username}' 已被其他用户使用")
                        existing_user.username = username  # type: ignore[assignment]
                    
                    await db.commit()
                    await db.refresh(existing_user)
                    
                    results.append(ImportUserResponse(
                        row_number=i,
                        student_id=student_id,
                        full_name=full_name,
                        status="success",
                        message="用户信息已更新",
                        user_id=existing_user.id  # type: ignore[arg-type]
                    ))
                    updated_count += 1
                else:
                    # 创建新用户
                    # 检查用户名唯一性
                    if username:
                        username_check = select(User).where(
                            User.username == username,
                            User.is_deleted == False
                        )
                        username_result = await db.execute(username_check)
                        if username_result.scalar_one_or_none():
                            raise ValueError(f"用户名 '{username}' 已存在")
                    
                    new_user = User(
                        student_id=student_id,
                        username=username,
                        full_name=full_name,
                        class_name=class_name,
                        study_year=study_year,
                        role_code=role_code,
                        is_active=is_active
                    )
                    
                    db.add(new_user)
                    await db.commit()
                    await db.refresh(new_user)
                    
                    results.append(ImportUserResponse(
                        row_number=i,
                        student_id=student_id,
                        full_name=full_name,
                        status="success",
                        message="用户创建成功",
                        user_id=new_user.id  # type: ignore[arg-type]
                    ))
                    imported_count += 1
                    
            except Exception as e:
                error_msg = str(e)
                results.append(ImportUserResponse(
                    row_number=i,
                    student_id=row.get('学号', ''),
                    full_name=row.get('姓名', ''),
                    status="error",
                    message=error_msg
                ))
                error_count += 1
                # 回滚当前事务（如果有）
                await db.rollback()
        
        # 最终提交
        await db.commit()
        
        return UserImportResult(
            success=True if error_count == 0 else False,
            message=f"导入完成。成功导入: {imported_count}, 更新: {updated_count}, 失败: {error_count}",
            total_rows=len(results),
            imported_count=imported_count,
            updated_count=updated_count,
            error_count=error_count,
            errors=[r for r in results if r.status == "error"]
        )
        
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"导入用户失败: {str(e)}"
        )
