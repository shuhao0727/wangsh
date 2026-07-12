"""用户 CSV/XLSX 模板、解析和批量导入实现。"""

import csv
import io
from typing import Any, Dict, Iterator, List, Optional, Tuple
from urllib.parse import quote

from fastapi import HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import User
from app.utils.errors import safe_error_detail

from .policy import assert_role_assignment_allowed, assert_users_mutable
from .schemas import ImportUserResponse, UserImportResult

USER_IMPORT_HEADERS = ["学号", "姓名", "学年", "班级", "状态", "用户名", "角色"]
USER_IMPORT_REQUIRED_FIELDS = ["学号", "姓名"]
USER_IMPORT_TEMPLATE_ROWS = [
    ["20230001", "张三", "2025", "高一(1)班", "true", "zhangsan", "student"],
    ["20230002", "李四", "2025", "高一(2)班", "true", "lisi", "student"],
    ["20230003", "王五", "2025", "高一(3)班", "false", "wangwu", "student"],
    ["20230004", "赵六", "2025", "高一(1)班", "true", "", "teacher"],
]


def normalize_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value).strip()


def parse_csv_rows(content: bytes) -> Tuple[List[str], Iterator[Dict[str, str]]]:
    content_text = content.decode("utf-8-sig")
    csv_file = io.StringIO(content_text)
    reader = csv.DictReader(csv_file)
    if not reader.fieldnames:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="CSV文件为空或格式不正确",
        )

    fieldnames = [str(name).strip() for name in reader.fieldnames if name]
    return fieldnames, reader


def parse_xlsx_rows(content: bytes) -> Tuple[List[str], Iterator[Dict[str, str]]]:
    workbook = load_workbook(filename=io.BytesIO(content), data_only=True, read_only=True)
    worksheet = workbook.active
    assert worksheet is not None  # XLSX 文件总是有 active sheet
    row_iterator = worksheet.iter_rows(values_only=True)
    header_row = next(row_iterator, None)
    if not header_row:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel文件为空或格式不正确",
        )

    fieldnames = [normalize_cell_value(value) for value in header_row]
    if not any(fieldnames):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel文件缺少表头",
        )

    def row_generator() -> Iterator[Dict[str, str]]:
        for row in row_iterator:
            current_row = row or ()
            parsed_row: Dict[str, str] = {}
            for idx, header in enumerate(fieldnames):
                if not header:
                    continue
                value = current_row[idx] if idx < len(current_row) else ""
                parsed_row[header] = normalize_cell_value(value)
            yield parsed_row

    return [header for header in fieldnames if header], row_generator()


def parse_import_rows(
    file_name: str,
    content: bytes,
) -> Tuple[List[str], Iterator[Dict[str, str]]]:
    lower_name = file_name.lower()
    if lower_name.endswith((".csv", ".txt")):
        return parse_csv_rows(content)
    if lower_name.endswith(".xlsx"):
        return parse_xlsx_rows(content)
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="只支持 CSV、TXT 或 XLSX 文件",
    )


def build_user_import_template(format: str) -> StreamingResponse:
    file_name = f"user_import_template.{format}"
    encoded_name = quote(file_name)
    headers = {"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_name}"}

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(USER_IMPORT_HEADERS)
        writer.writerow(["# 注意：角色列可选值：student(学生)/teacher(教师)/admin(管理员)/super_admin(超级管理员)"])
        writer.writerow(["# 状态：true=激活，false=未激活（可选，默认为true）"])
        writer.writerow(["# 用户名：可选字段，如果不填将使用学号作为登录账号"])
        writer.writerow(["# 示例数据："])
        for sample_row in USER_IMPORT_TEMPLATE_ROWS:
            writer.writerow(sample_row)

        csv_bytes = output.getvalue().encode("utf-8-sig")
        return StreamingResponse(
            io.BytesIO(csv_bytes),
            media_type="text/csv; charset=utf-8",
            headers=headers,
        )

    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None  # Workbook() 总是有一个 active sheet
    worksheet.title = "用户导入模板"
    worksheet.append(USER_IMPORT_HEADERS)
    for sample_row in USER_IMPORT_TEMPLATE_ROWS:
        worksheet.append(sample_row)

    notes_sheet = workbook.create_sheet(title="填写说明")
    notes_sheet.append(["说明"])
    notes_sheet.append(["1. 角色列可选值：student(学生)/teacher(教师)/admin(管理员)/super_admin(超级管理员)"])
    notes_sheet.append(["2. 状态字段可填 true/false（默认为 true）"])
    notes_sheet.append(["3. 用户名字段可选，不填将使用学号作为登录账号"])
    notes_sheet.append(["4. 学号、姓名为必填字段"])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


def _parse_import_user(row: Dict[str, str]) -> Dict[str, Any]:
    student_id = row.get("学号", "").strip()
    full_name = row.get("姓名", "").strip()
    if not student_id or not full_name:
        raise ValueError("学号和姓名为必填字段")

    status_str = row.get("状态", "true").strip().lower()
    if status_str in ["false", "0", "no", "否"]:
        is_active = False
    elif status_str in ["true", "1", "yes", "是", ""]:
        is_active = True
    else:
        raise ValueError(f"状态值无效: {status_str}")

    role_code = row.get("角色", "student").strip() or "student"
    if role_code not in ("student", "teacher", "admin", "super_admin"):
        role_code = "student"

    return {
        "student_id": student_id,
        "full_name": full_name,
        "study_year": row.get("学年", "").strip() or None,
        "class_name": row.get("班级", "").strip() or None,
        "username": row.get("用户名", "").strip() or None,
        "role_code": role_code,
        "is_active": is_active,
    }


async def _find_import_user(db: AsyncSession, student_id: str) -> User | None:
    query = select(User).where(
        User.student_id == student_id,
        User.is_deleted == False,
    ).with_for_update()
    result = await db.execute(query)
    return result.scalar_one_or_none()


async def _ensure_username_available(
    db: AsyncSession,
    username: Optional[str],
    *,
    excluded_user_id: Optional[int] = None,
) -> None:
    if not username:
        return
    conditions = [User.username == username, User.is_deleted == False]
    if excluded_user_id is not None:
        conditions.append(User.id != excluded_user_id)
    result = await db.execute(select(User).where(*conditions))
    if result.scalar_one_or_none():
        suffix = "已被其他用户使用" if excluded_user_id is not None else "已存在"
        raise ValueError(f"用户名 '{username}' {suffix}")


async def _update_import_user(
    db: AsyncSession,
    user: User,
    data: Dict[str, Any],
    row_number: int,
) -> ImportUserResponse:
    assert_users_mutable(data["current_user"], [user])
    user.full_name = data["full_name"]  # type: ignore[assignment]
    user.study_year = data["study_year"] or user.study_year  # type: ignore[assignment]
    user.class_name = data["class_name"] or user.class_name  # type: ignore[assignment]
    user.is_active = data["is_active"]  # type: ignore[assignment]
    username = data["username"]
    if username and username != user.username:
        await _ensure_username_available(
            db,
            username,
            excluded_user_id=user.id,
        )
        user.username = username  # type: ignore[assignment]
    await db.flush()
    return ImportUserResponse(
        row_number=row_number,
        student_id=data["student_id"],
        full_name=data["full_name"],
        status="success",
        message="用户信息已更新",
        user_id=user.id,
    )


async def _create_import_user(
    db: AsyncSession,
    data: Dict[str, Any],
    row_number: int,
) -> ImportUserResponse:
    await _ensure_username_available(db, data["username"])
    user = User(
        student_id=data["student_id"],
        username=data["username"],
        full_name=data["full_name"],
        class_name=data["class_name"],
        study_year=data["study_year"],
        role_code=data["role_code"],
        is_active=data["is_active"],
    )
    db.add(user)
    await db.flush()
    return ImportUserResponse(
        row_number=row_number,
        student_id=data["student_id"],
        full_name=data["full_name"],
        status="success",
        message="用户创建成功",
        user_id=user.id,
    )


async def _import_user_row(
    db: AsyncSession,
    row: Dict[str, str],
    current_user: Dict[str, Any],
    row_number: int,
) -> tuple[ImportUserResponse, bool]:
    data = _parse_import_user(row)
    data["current_user"] = current_user
    assert_role_assignment_allowed(current_user, data["role_code"])
    existing_user = await _find_import_user(db, data["student_id"])
    if existing_user:
        return await _update_import_user(db, existing_user, data, row_number), False
    return await _create_import_user(db, data, row_number), True


def _validate_import_headers(fieldnames: List[str]) -> None:
    for field in USER_IMPORT_REQUIRED_FIELDS:
        if field not in fieldnames:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"导入文件缺少必要字段: {field}",
            )


async def _process_import_rows(
    db: AsyncSession,
    reader: Iterator[Dict[str, str]],
    current_user: Dict[str, Any],
) -> tuple[List[ImportUserResponse], int, int]:
    results: List[ImportUserResponse] = []
    imported_count = 0
    updated_count = 0
    for row_number, row in enumerate(reader, start=1):
        if any(str(value).startswith("#") for value in row.values() if value):
            continue
        savepoint = None
        try:
            savepoint = await db.begin_nested()
            result, created = await _import_user_row(
                db,
                row,
                current_user,
                row_number,
            )
            results.append(result)
            imported_count += int(created)
            updated_count += int(not created)
        except Exception as exc:
            if savepoint is not None:
                await savepoint.rollback()
            results.append(
                ImportUserResponse(
                    row_number=row_number,
                    student_id=row.get("学号", ""),
                    full_name=row.get("姓名", ""),
                    status="error",
                    message=str(exc),
                )
            )
    return results, imported_count, updated_count


async def import_users(
    file: UploadFile,
    current_user: Dict[str, Any],
    db: AsyncSession,
) -> UserImportResult:
    try:
        if not file.filename:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="文件名不能为空",
            )
        fieldnames, reader = parse_import_rows(file.filename, await file.read())
        _validate_import_headers(fieldnames)
        results, imported_count, updated_count = await _process_import_rows(
            db,
            reader,
            current_user,
        )
        await db.commit()
        errors = [result for result in results if result.status == "error"]
        error_count = len(errors)
        return UserImportResult(
            success=error_count == 0,
            message=f"导入完成。成功导入: {imported_count}, 更新: {updated_count}, 失败: {error_count}",
            total_rows=len(results),
            imported_count=imported_count,
            updated_count=updated_count,
            error_count=error_count,
            errors=errors,
        )
    except HTTPException:
        raise
    except Exception as exc:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=safe_error_detail("导入用户失败", exc),
        )
