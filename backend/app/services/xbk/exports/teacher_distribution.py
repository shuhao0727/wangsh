from __future__ import annotations

from io import BytesIO
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import Integer, and_, case, cast, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import XbkCourse, XbkSelection, XbkStudent
from app.services.xbk.exports.common import THIN_BORDER, class_sort_key, safe_sheet_name


def _title_year_range(year: int, year_start: Optional[int], year_end: Optional[int]) -> Tuple[int, int]:
    ys = year_start if year_start is not None else year
    ye = year_end if year_end is not None else year + 1
    return ys, ye


def _guess_grade(class_name: Optional[str]) -> str:
    if not class_name:
        return ""
    s = str(class_name)
    for g in ["高一", "高二", "高三", "初一", "初二", "初三"]:
        if s.startswith(g):
            return g
    return ""


def _cn_len(value: object) -> int:
    if value is None:
        return 0
    s = str(value)
    n = 0
    for ch in s:
        n += 2 if ord(ch) > 127 else 1
    return n


def _apply_excel_style(ws, header_fill_color: str = "CCE5FF") -> None:
    header_fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
    header_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    merged_ranges = list(ws.merged_cells.ranges) if ws.merged_cells else []

    def is_top_left_of_merged(cell) -> bool:
        for merged_range in merged_ranges:
            if cell.row == merged_range.min_row and cell.column == merged_range.min_col:
                return True
        return False

    def is_in_merged(cell) -> bool:
        for merged_range in merged_ranges:
            if merged_range.min_row <= cell.row <= merged_range.max_row and merged_range.min_col <= cell.column <= merged_range.max_col:
                return True
        return False

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        if not is_in_merged(cell) or is_top_left_of_merged(cell):
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if not is_in_merged(cell) or is_top_left_of_merged(cell):
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")


def _apply_merged_cell_borders(ws) -> None:
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    if not ws.merged_cells:
        return
    for merged_range in list(ws.merged_cells.ranges):
        min_row = merged_range.min_row
        max_row = merged_range.max_row
        min_col = merged_range.min_col
        max_col = merged_range.max_col

        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=min_row, column=col)
            cell.border = Border(
                left=cell.border.left if cell.border else Side(style="thin"),
                right=cell.border.right if cell.border else Side(style="thin"),
                top=Side(style="thin"),
                bottom=cell.border.bottom if cell.border else Side(style="thin"),
            )
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=max_row, column=col)
            cell.border = Border(
                left=cell.border.left if cell.border else Side(style="thin"),
                right=cell.border.right if cell.border else Side(style="thin"),
                top=cell.border.top if cell.border else Side(style="thin"),
                bottom=Side(style="thin"),
            )
        for row in range(min_row, max_row + 1):
            cell = ws.cell(row=row, column=min_col)
            cell.border = Border(
                left=Side(style="thin"),
                right=cell.border.right if cell.border else Side(style="thin"),
                top=cell.border.top if cell.border else Side(style="thin"),
                bottom=cell.border.bottom if cell.border else Side(style="thin"),
            )
        for row in range(min_row, max_row + 1):
            cell = ws.cell(row=row, column=max_col)
            cell.border = Border(
                left=cell.border.left if cell.border else Side(style="thin"),
                right=Side(style="thin"),
                top=cell.border.top if cell.border else Side(style="thin"),
                bottom=cell.border.bottom if cell.border else Side(style="thin"),
            )

        ws.cell(row=min_row, column=min_col).border = thin_border
        ws.cell(row=min_row, column=max_col).border = thin_border
        ws.cell(row=max_row, column=min_col).border = thin_border
        ws.cell(row=max_row, column=max_col).border = thin_border


def _adjust_column_widths(ws) -> None:
    recommended_widths = {"A": 12, "B": 12, "C": 10, "D": 10, "E": 10, "F": 10, "G": 10, "H": 10, "I": 10}
    max_widths = {"A": 15, "B": 15, "C": 12, "D": 12, "E": 12, "F": 12, "G": 12, "H": 12, "I": 12}
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        recommended_width = recommended_widths.get(col_letter, 10)
        max_width = max_widths.get(col_letter, 15)
        max_content_length = 0
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            max_content_length = max(max_content_length, _cn_len(v))
        content_based_width = max_content_length + 2
        final_width = max(recommended_width, min(content_based_width, max_width))
        ws.column_dimensions[col_letter].width = final_width


def _adjust_row_heights(ws) -> None:
    if ws.max_row >= 1:
        ws.row_dimensions[1].height = 35
    if ws.max_row >= 2:
        text = ws.cell(row=2, column=1).value
        total_width = 0
        for letter in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
            w = ws.column_dimensions[letter].width
            total_width += int(w) if w else 10
        per_line = max(30, int(total_width * 1.2))
        lines = 1
        if text:
            lines = max(1, (_cn_len(text) + per_line - 1) // per_line)
        ws.row_dimensions[2].height = min(25 + (lines - 1) * 12, 60)
    if ws.max_row >= 3:
        ws.row_dimensions[3].height = 20
    if ws.max_row >= 4:
        ws.row_dimensions[4].height = 20

    for row in range(5, ws.max_row + 1):
        max_lines = 1
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is None:
                continue
            text = str(cell.value)
            lines = text.count("\n") + 1
            if len(text) > 30:
                lines += len(text) // 30
            max_lines = max(max_lines, lines)
        base_height = 18
        extra_height = (max_lines - 1) * 5
        ws.row_dimensions[row].height = min(base_height + extra_height, 40)


def _set_page_settings(ws) -> None:
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2
    ws.print_options.gridLines = True
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = "1:5"
    ws.sheet_properties.pageSetUpPr.fitToPage = True


async def build_teacher_distribution_xlsx(
    db: AsyncSession,
    year: int,
    term: str,
    class_name: Optional[str],
    year_start: Optional[int],
    year_end: Optional[int],
) -> BytesIO:
    ys, ye = _title_year_range(year, year_start, year_end)

    numeric_course_code = case(
        (XbkCourse.course_code.op("~")("^[0-9]+$"), cast(XbkCourse.course_code, Integer)),
        else_=None,
    )
    course_stmt = (
        select(XbkCourse)
        .where(XbkCourse.is_deleted.is_(False), XbkCourse.year == year, XbkCourse.term == term)
        .order_by(numeric_course_code.asc().nulls_last(), XbkCourse.course_code.asc())
    )
    courses = (await db.execute(course_stmt)).scalars().all()

    wb = Workbook()
    wb.remove(wb.active)

    for c in courses:
        stmt = (
            select(
                XbkStudent.class_name.label("class_name"),
                func.coalesce(XbkSelection.name, XbkStudent.name).label("student_name"),
            )
            .select_from(XbkSelection)
            .outerjoin(
                XbkStudent,
                and_(
                    XbkStudent.is_deleted.is_(False),
                    XbkStudent.year == XbkSelection.year,
                    XbkStudent.term == XbkSelection.term,
                    XbkStudent.student_no == XbkSelection.student_no,
                ),
            )
            .where(
                XbkSelection.is_deleted.is_(False),
                XbkSelection.year == year,
                XbkSelection.term == term,
                XbkSelection.course_code == c.course_code,
            )
            .order_by(XbkStudent.class_name.asc(), XbkStudent.student_no.asc())
        )
        if class_name:
            stmt = stmt.where(XbkStudent.class_name == class_name)
        rows = (await db.execute(stmt)).all()
        if not rows:
            continue
        rows = sorted(rows, key=lambda r: (class_sort_key(str(r.class_name or "")), str(r.student_name or "")))

        grade = _guess_grade(str(rows[0].class_name)) if rows and rows[0].class_name else ""
        grade = grade or "年级"
        sheet_name = safe_sheet_name(str(c.course_code))
        ws = wb.create_sheet(sheet_name)

        ws.merge_cells("A1:I1")
        ws["A1"] = f"{ys}-{ye} 学年江苏省昆山中学 {grade} 校本课程学生签到表"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35

        ws.merge_cells("A2:I2")
        ws["A2"] = (
            f"课程代码: {c.course_code}  课程名称: {c.course_name or ''}  "
            f"课程负责人: {c.teacher or ''}  上课地点: {c.location or ''}  人数: {len(rows)}"
        )
        ws["A2"].font = Font(size=12, bold=True)
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 25

        ws.merge_cells("A3:B4")
        ws["A3"] = "任课教师签名"
        ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A3"].font = Font(bold=True)

        for idx, text in enumerate(["一", "二", "三", "四", "五", "六", "七"], start=3):
            cell = ws.cell(row=3, column=idx)
            cell.value = f"第{text}次"
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)

        ws.cell(row=5, column=1).value = "班级"
        ws.cell(row=5, column=2).value = "姓名"
        ws.cell(row=5, column=1).font = Font(bold=True)
        ws.cell(row=5, column=2).font = Font(bold=True)
        ws.cell(row=5, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=5, column=2).alignment = Alignment(horizontal="center", vertical="center")
        for col in range(3, 10):
            ws.cell(row=5, column=col).value = None

        start_row = 6
        for idx, r in enumerate(rows):
            row_no = start_row + idx
            ws.cell(row=row_no, column=1).value = r.class_name
            ws.cell(row=row_no, column=2).value = r.student_name

        ws.cell(row=1, column=1).fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        ws.cell(row=2, column=1).fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
        ws.cell(row=3, column=1).fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        for col in range(3, 10):
            ws.cell(row=3, column=col).fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

        _apply_excel_style(ws)
        _apply_merged_cell_borders(ws)
        _adjust_column_widths(ws)
        _adjust_row_heights(ws)
        ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        _set_page_settings(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
