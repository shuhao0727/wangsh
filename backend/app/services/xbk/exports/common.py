from __future__ import annotations

import re
from typing import Iterable, Optional, Sequence, Tuple

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


THIN_SIDE = Side(style="thin", color="D9D9D9")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


def cn_len(value: object) -> int:
    if value is None:
        return 0
    s = str(value)
    n = 0
    for ch in s:
        n += 2 if ord(ch) > 127 else 1
    return n


def auto_adjust_column_width(ws, min_width: int = 10, max_width: int = 48, sample_rows: int = 300) -> None:
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < 1 or max_col < 1:
        return
    for col in range(1, max_col + 1):
        max_len = 0
        for row in range(1, min(max_row, sample_rows) + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, cn_len(v))
        width = min(max(int(max_len * 1.1) + 2, min_width), max_width)
        ws.column_dimensions[get_column_letter(col)].width = width


def apply_table_style(
    ws,
    header_row: int,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
    header_fill_color: str = "CCE5FF",
    zebra_fill_color: str = "F5F5F5",
) -> None:
    header_fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
    header_font = Font(bold=True, size=11, color="000000")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    zebra_fill = PatternFill(start_color=zebra_fill_color, end_color=zebra_fill_color, fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for col in range(start_col, end_col + 1):
        c = ws.cell(row=header_row, column=col)
        c.fill = header_fill
        c.font = header_font
        c.border = THIN_BORDER
        c.alignment = header_align

    for r in range(start_row, end_row + 1):
        fill = zebra_fill if r % 2 == 0 else white_fill
        for col in range(start_col, end_col + 1):
            c = ws.cell(row=r, column=col)
            c.border = THIN_BORDER
            c.alignment = body_align
            c.fill = fill


def set_print_a4_landscape(ws, print_title_rows: Optional[Tuple[int, int]] = None) -> None:
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2
    ws.print_options.horizontalCentered = True
    ws.print_options.gridLines = False
    if print_title_rows:
        ws.print_title_rows = f"{print_title_rows[0]}:{print_title_rows[1]}"


def safe_sheet_name(name: str) -> str:
    s = (name or "").strip()
    s = s.replace("/", " ").replace("\\", " ").replace(":", " ").replace("*", " ").replace("?", " ").replace("[", "(").replace("]", ")")
    s = s[:31]
    return s if s else "Sheet"


def class_sort_key(name: str):
    s = (name or "").strip()
    if not s:
        return (9, "", 9999, s)
    m = re.match(r"^(高一|高二|高三|初一|初二|初三)\s*[\(（]?(\d+)[\)）]?\s*班?$", s)
    if m:
        grade = m.group(1)
        num = int(m.group(2))
        grade_order = {"初一": 1, "初二": 2, "初三": 3, "高一": 4, "高二": 5, "高三": 6}.get(grade, 9)
        return (0, grade_order, num, s)
    m2 = re.search(r"(\d+)", s)
    if m2:
        return (1, s[: m2.start()], int(m2.group(1)), s)
    return (2, s, 9999, s)
