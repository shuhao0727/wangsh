from __future__ import annotations

from io import BytesIO
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import Integer, and_, case, cast, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import XbkCourse, XbkSelection, XbkStudent
from app.services.xbk.exports.common import class_sort_key, cn_len, safe_sheet_name


def _title_year_range(year: int, year_start: Optional[int], year_end: Optional[int]) -> Tuple[int, int]:
    ys = year_start if year_start is not None else year
    ye = year_end if year_end is not None else year + 1
    return ys, ye


def _guess_grade(class_name: Optional[str]) -> str:
    if not class_name:
        return ""
    s = str(class_name)
    for g in ["高一", "高二", "高三", "初一", "初二", "初三"]:
        if s.startswith(g):
            return g
    return ""


def _apply_distribution_style(ws) -> None:
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    alternate_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if cell.row == 2:
                cell.fill = header_fill
                cell.font = Font(bold=True, size=11)
            elif cell.row >= 3:
                if cell.row % 2 == 0:
                    cell.fill = alternate_fill


def _apply_merged_cell_borders(ws) -> None:
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    if not ws.merged_cells:
        return
    for merged_range in list(ws.merged_cells.ranges):
        min_row = merged_range.min_row
        max_row = merged_range.max_row
        min_col = merged_range.min_col
        max_col = merged_range.max_col
        for col in range(min_col, max_col + 1):
            ws.cell(row=min_row, column=col).border = thin_border
            ws.cell(row=max_row, column=col).border = thin_border
        for row in range(min_row, max_row + 1):
            ws.cell(row=row, column=min_col).border = thin_border
            ws.cell(row=row, column=max_col).border = thin_border
        ws.cell(row=min_row, column=min_col).border = thin_border
        ws.cell(row=min_row, column=max_col).border = thin_border
        ws.cell(row=max_row, column=min_col).border = thin_border
        ws.cell(row=max_row, column=max_col).border = thin_border


def _adjust_column_widths(ws) -> None:
    recommended_widths = {"A": 8, "B": 25, "C": 15, "D": 18, "E": 10}
    max_widths = {"A": 12, "B": 30, "C": 20, "D": 25, "E": 15}
    for letter, col_idx in zip(["A", "B", "C", "D", "E"], range(1, 6)):
        recommended_width = recommended_widths[letter]
        max_width = max_widths[letter]
        max_content_len = cn_len(ws.cell(row=1, column=col_idx).value)
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col_idx).value
            if v is None:
                continue
            max_content_len = max(max_content_len, cn_len(v))
        content_based = max_content_len + 2
        ws.column_dimensions[letter].width = max(recommended_width, min(content_based, max_width))


def _adjust_row_heights(ws) -> None:
    if ws.max_row >= 1:
        ws.row_dimensions[1].height = 35
    if ws.max_row >= 2:
        ws.row_dimensions[2].height = 25
    for row in range(3, ws.max_row + 1):
        max_lines = 1
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            text = str(v)
            lines = text.count("\n") + 1
            if len(text) > 30:
                lines += len(text) // 30
            max_lines = max(max_lines, lines)
        base_height = 18
        extra_height = (max_lines - 1) * 5
        ws.row_dimensions[row].height = min(base_height + extra_height, 40)


def _set_page_settings(ws) -> None:
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2
    ws.print_options.gridLines = True
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = "1:2"
    ws.sheet_properties.pageSetUpPr.fitToPage = True


async def build_class_distribution_xlsx(
    db: AsyncSession,
    year: int,
    term: str,
    class_name: Optional[str],
    year_start: Optional[int],
    year_end: Optional[int],
) -> BytesIO:
    ys, ye = _title_year_range(year, year_start, year_end)

    numeric_course_code = case(
        (XbkSelection.course_code.op("~")("^[0-9]+$"), cast(XbkSelection.course_code, Integer)),
        else_=None,
    )
    stmt = (
        select(
            XbkStudent.class_name.label("class_name"),
            XbkStudent.student_no.label("student_no"),
            func.coalesce(XbkSelection.name, XbkStudent.name).label("student_name"),
            XbkSelection.course_code.label("course_code"),
            XbkCourse.course_name.label("course_name"),
            XbkCourse.teacher.label("teacher"),
            XbkCourse.location.label("location"),
        )
        .select_from(XbkSelection)
        .outerjoin(
            XbkStudent,
            and_(
                XbkStudent.is_deleted.is_(False),
                XbkStudent.year == XbkSelection.year,
                XbkStudent.term == XbkSelection.term,
                XbkStudent.student_no == XbkSelection.student_no,
            ),
        )
        .outerjoin(
            XbkCourse,
            and_(
                XbkCourse.is_deleted.is_(False),
                XbkCourse.year == XbkSelection.year,
                XbkCourse.term == XbkSelection.term,
                XbkCourse.course_code == XbkSelection.course_code,
            ),
        )
        .where(XbkSelection.is_deleted.is_(False), XbkSelection.year == year, XbkSelection.term == term)
        .order_by(
            XbkStudent.class_name.asc(),
            numeric_course_code.asc().nulls_last(),
            XbkSelection.course_code.asc(),
            XbkStudent.student_no.asc(),
        )
    )
    if class_name:
        stmt = stmt.where(XbkStudent.class_name == class_name)
    rows = (await db.execute(stmt)).all()

    grouped: Dict[str, List[dict]] = {}
    for r in rows:
        grouped.setdefault(str(r.class_name or "未知班级"), []).append(
            {
                "course_code": r.course_code,
                "course_name": r.course_name,
                "teacher": r.teacher,
                "location": r.location,
                "student_name": r.student_name,
            }
        )

    def _course_code_key(v: object):
        s = "" if v is None else str(v).strip()
        if s.isdigit():
            return (0, int(s), s)
        return (1, s, s)

    wb = Workbook()
    wb.remove(wb.active)

    for cls in sorted(grouped.keys(), key=class_sort_key):
        items = grouped[cls]
        grade = _guess_grade(cls) or "年级"
        ws = wb.create_sheet(safe_sheet_name(cls))
        ws.merge_cells("A1:E1")
        ws["A1"] = f"{ys}-{ye} 学年江苏省昆山中学 {grade} {cls}班选课分发表"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35

        ws.append(["课程代码", "课程名称", "课程负责人", "上课地点", "姓名"])
        ws.row_dimensions[2].height = 25
        items_sorted = sorted(items, key=lambda it: (_course_code_key(it.get("course_code")), str(it.get("student_name") or "")))
        for it in items_sorted:
            ws.append([it["course_code"], it["course_name"], it["teacher"], it["location"], it["student_name"]])

        _apply_distribution_style(ws)
        _apply_merged_cell_borders(ws)
        _adjust_column_widths(ws)
        _adjust_row_heights(ws)
        _set_page_settings(ws)
        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:E{ws.max_row}"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
