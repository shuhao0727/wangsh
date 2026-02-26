from __future__ import annotations

from io import BytesIO
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import Integer, case, cast, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import XbkCourse, XbkStudent
from app.services.xbk.exports.common import apply_table_style, auto_adjust_column_width, class_sort_key, safe_sheet_name


def _title_year_range(year: int, year_start: Optional[int], year_end: Optional[int]) -> Tuple[int, int]:
    ys = year_start if year_start is not None else year
    ye = year_end if year_end is not None else year + 1
    return ys, ye


def _parse_int(value: object) -> Optional[int]:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    try:
        return int(float(s))
    except Exception:
        return None


def _get_course_code_limits(ws) -> List[str]:
    codes: List[str] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row:
            continue
        code = row[0]
        if code is None:
            continue
        codes.append(str(code))
    return codes


def _set_course_code_validation(ws, column_letter: str, valid_course_codes: List[str], catalog_sheet_name: str) -> None:
    int_codes = []
    for code in valid_course_codes:
        v = _parse_int(code)
        if v is not None:
            int_codes.append(v)
    min_code = min(int_codes) if int_codes else None
    max_code = max(int_codes) if int_codes else None
    for row in range(2, ws.max_row + 1):
        if min_code is not None and max_code is not None:
            formula = (
                f"=AND(ISNUMBER({column_letter}{row}),"
                f"{column_letter}{row}>={min_code},"
                f"{column_letter}{row}<={max_code},"
                f"COUNTIF({column_letter}:{column_letter},{column_letter}{row})<=VLOOKUP({column_letter}{row},'{catalog_sheet_name}'!A:E,4,0))"
            )
        else:
            formula = f"=ISNUMBER({column_letter}{row})"
        dv = DataValidation(
            type="custom",
            formula1=formula,
            showErrorMessage=True,
            errorTitle="错误",
            error="只能输入有效的课程代码，并且数量不能超过限制",
        )
        ws.add_data_validation(dv)
        dv.add(f"{column_letter}{row}")


def _lock_sheet(ws, unlock_header: Optional[str] = None, unlock_ranges: Optional[List[str]] = None) -> None:
    ws.protection.sheet = True
    ws.protection.set_password("")
    for row in ws.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True)
    if unlock_header:
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == unlock_header:
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=col).protection = Protection(locked=False)
    if unlock_ranges:
        for rng in unlock_ranges:
            for row in ws[rng]:
                for cell in row:
                    cell.protection = Protection(locked=False)


def _cn_len(value: object) -> int:
    if value is None:
        return 0
    s = str(value)
    n = 0
    for ch in s:
        n += 2 if ord(ch) > 127 else 1
    return n


def _clamp(v: int, mn: int, mx: int) -> int:
    return max(mn, min(mx, v))


def _adjust_catalog_dimensions(ws) -> None:
    col_specs = [
        ("A", 1, "课程代码", 10, 14),
        ("B", 2, "课程名称", 20, 44),
        ("C", 3, "课程负责人", 12, 22),
        ("D", 4, "各班限报人数", 10, 14),
        ("E", 5, "上课地点", 16, 46),
    ]

    for letter, idx, title, min_w, max_w in col_specs:
        max_len = _cn_len(title)
        for r in range(2, min(ws.max_row, 300) + 1):
            v = ws.cell(row=r, column=idx).value
            if v is None:
                continue
            max_len = max(max_len, _cn_len(v))
        width = _clamp(int(max_len * 0.9) + 2, min_w, max_w)
        ws.column_dimensions[letter].width = width

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 20
    for r in range(3, ws.max_row + 1):
        max_lines = 1
        for letter, idx, _, _, _ in col_specs:
            v = ws.cell(row=r, column=idx).value
            if v is None:
                continue
            text = str(v).strip()
            if not text:
                continue
            base_lines = text.count("\n") + 1
            width = ws.column_dimensions[letter].width or 10
            approx_per_line = max(6, int(width * 1.1))
            wrapped_lines = max(base_lines, (_cn_len(text) + approx_per_line - 1) // approx_per_line)
            max_lines = max(max_lines, wrapped_lines)
        ws.row_dimensions[r].height = min(18 * max_lines, 72)


async def build_student_course_selection_xlsx(
    db: AsyncSession,
    year: int,
    term: str,
    class_name: Optional[str],
    year_start: Optional[int],
    year_end: Optional[int],
) -> BytesIO:
    courses = (
        await db.execute(
            select(XbkCourse)
            .where(XbkCourse.is_deleted.is_(False), XbkCourse.year == year, XbkCourse.term == term)
            .order_by(
                case(
                    (XbkCourse.course_code.op("~")("^[0-9]+$"), cast(XbkCourse.course_code, Integer)),
                    else_=None,
                )
                .asc()
                .nulls_last(),
                XbkCourse.course_code.asc(),
            )
        )
    ).scalars().all()

    stu_stmt = select(XbkStudent).where(
        XbkStudent.is_deleted.is_(False),
        XbkStudent.year == year,
        XbkStudent.term == term,
    )
    if class_name:
        stu_stmt = stu_stmt.where(XbkStudent.class_name == class_name)
    students = (await db.execute(stu_stmt.order_by(XbkStudent.class_name.asc(), XbkStudent.student_no.asc()))).scalars().all()

    students_by_class: Dict[str, List[XbkStudent]] = {}
    for s in students:
        students_by_class.setdefault(s.class_name, []).append(s)

    wb = Workbook()
    wb.remove(wb.active)

    catalog = wb.create_sheet("校本课程目录")
    catalog.append(["课程代码", "课程名称", "课程负责人", "各班限报人数", "上课地点"])
    for c in courses:
        code = _parse_int(c.course_code) if _parse_int(c.course_code) is not None else c.course_code
        catalog.append([code, c.course_name, c.teacher, int(c.quota or 0), c.location])

    ys, ye = _title_year_range(year, year_start, year_end)

    catalog.insert_rows(1)
    catalog["A1"] = f"江苏省昆山中学校本课程目录（{ys}-{ye} {term}）"
    catalog["A1"].font = Font(size=14, bold=True)
    catalog["A1"].alignment = Alignment(horizontal="center", vertical="center")
    catalog.merge_cells("A1:E1")

    apply_table_style(catalog, header_row=2, start_row=3, end_row=catalog.max_row, start_col=1, end_col=5)
    _adjust_catalog_dimensions(catalog)
    catalog.freeze_panes = "A3"
    catalog.auto_filter.ref = None
    _lock_sheet(catalog)

    valid_course_codes = _get_course_code_limits(catalog)

    for cls in sorted(students_by_class.keys(), key=class_sort_key):
        stus = students_by_class[cls]
        ws = wb.create_sheet(safe_sheet_name(cls))
        ws.append(["班级", "学号", "姓名", "课程代码"])
        for s in stus:
            ws.append([s.class_name, s.student_no, s.name, None])

        apply_table_style(ws, header_row=1, start_row=2, end_row=ws.max_row, start_col=1, end_col=4)
        auto_adjust_column_width(ws, max_width=30)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = None

        course_col_letter = get_column_letter(4)
        _set_course_code_validation(ws, course_col_letter, valid_course_codes, "校本课程目录")
        _lock_sheet(ws, unlock_header="课程代码")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
