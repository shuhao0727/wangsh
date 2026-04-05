#!/usr/bin/env python3
from __future__ import annotations

import argparse
import io
import json
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook

if __package__ in (None, ""):
    import sys
    from pathlib import Path

    sys.path.append(str(Path(__file__).resolve().parents[2]))

from scripts.xbk.common import (
    auth_headers,
    dump_json,
    ensure_dir,
    get_http_client,
    get_settings,
    login_admin,
    utc_now_iso,
)


def _find_exact(items: List[Dict[str, Any]], key: str, value: str) -> Dict[str, Any]:
    for row in items:
        if str(row.get(key) or "") == value:
            return row
    raise RuntimeError(f"Cannot find row with {key}={value}")


def _list_items(
    client,
    api_base: str,
    token: str,
    endpoint: str,
    *,
    year: int,
    term: str,
    search_text: str = "",
) -> List[Dict[str, Any]]:
    resp = client.get(
        f"{api_base}{endpoint}",
        params={
            "page": 1,
            "size": 500,
            "year": year,
            "term": term,
            "search_text": search_text,
        },
        headers=auth_headers(token),
    )
    resp.raise_for_status()
    body = resp.json() or {}
    items = body.get("items") if isinstance(body, dict) else None
    if not isinstance(items, list):
        raise RuntimeError(f"Unexpected list response from {endpoint}: {body}")
    return [it for it in items if isinstance(it, dict)]


def _preview_import(client, api_base: str, token: str, scope: str, file_path: Path) -> Dict[str, Any]:
    with file_path.open("rb") as f:
        resp = client.post(
            f"{api_base}/xbk/import/preview",
            params={"scope": scope},
            files={"file": (file_path.name, f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=auth_headers(token),
        )
    resp.raise_for_status()
    body = resp.json() or {}
    if not isinstance(body, dict):
        raise RuntimeError(f"Unexpected preview response for {scope}: {body}")
    return body


def _import_data(client, api_base: str, token: str, scope: str, file_path: Path) -> Dict[str, Any]:
    with file_path.open("rb") as f:
        resp = client.post(
            f"{api_base}/xbk/import",
            params={"scope": scope, "skip_invalid": True},
            files={"file": (file_path.name, f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=auth_headers(token),
        )
    resp.raise_for_status()
    body = resp.json() or {}
    if not isinstance(body, dict):
        raise RuntimeError(f"Unexpected import response for {scope}: {body}")
    return body


def _assert_import_result(scope: str, result: Dict[str, Any], expected: Dict[str, Any]) -> None:
    for key in ["total_rows", "processed", "inserted", "updated", "invalid"]:
        got = int(result.get(key) or 0)
        want = int(expected.get(key) or 0)
        if got != want:
            raise RuntimeError(f"{scope} import {key} mismatch: got={got}, expected={want}")


def _validate_xlsx(content: bytes, required_headers: List[str]) -> Tuple[List[str], int]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    headers: List[str] = []
    for cell in ws[1]:
        headers.append(str(cell.value) if cell.value is not None else "")
    row_count = max(int(ws.max_row or 0) - 1, 0)
    for header in required_headers:
        if header not in headers:
            raise RuntimeError(f"Missing export header: {header}")
    if row_count <= 0:
        raise RuntimeError("Export row count is 0")
    return headers, row_count


def _delete_selection(client, api_base: str, token: str, selection_id: int) -> Dict[str, Any]:
    resp = client.delete(
        f"{api_base}/xbk/data/selections/{selection_id}",
        headers=auth_headers(token),
    )
    if resp.status_code != 204:
        raise RuntimeError(f"Delete selection failed: status={resp.status_code}, body={resp.text}")
    return {"status": resp.status_code, "selection_id": selection_id}


def run_smoke(manifest_path: Path | None = None) -> Dict[str, Any]:
    settings = get_settings()
    manifest_file = manifest_path or (settings.sample_dir / "manifest.json")
    if not manifest_file.exists():
        raise RuntimeError(f"Manifest not found: {manifest_file}")

    manifest = json.loads(manifest_file.read_text(encoding="utf-8"))
    year = int(manifest.get("year") or settings.year)
    term = str(manifest.get("term") or settings.term)

    report: Dict[str, Any] = {
        "timestamp": utc_now_iso(),
        "api_base": settings.api_base,
        "manifest": str(manifest_file),
        "year": year,
        "term": term,
        "steps": [],
        "status": "running",
    }

    def run_step(name: str, fn):
        try:
            data = fn()
            report["steps"].append({"name": name, "status": "passed", "data": data})
            return data
        except Exception as exc:  # noqa: BLE001
            report["steps"].append({"name": name, "status": "failed", "error": str(exc)})
            raise

    def checked_json(resp) -> Dict[str, Any]:
        resp.raise_for_status()
        body = resp.json() or {}
        if not isinstance(body, dict):
            raise RuntimeError(f"Expected JSON object, got: {body}")
        return body

    with get_http_client() as client:
        auth_token = login_admin(client, settings)
        run_step("admin-login", lambda: {"token_prefix": auth_token[:12]})

        expected = manifest.get("expected") or {}
        files = {k: Path(v) for k, v in (manifest.get("files") or {}).items()}

        for scope in ["students", "courses", "selections"]:
            if scope not in files:
                raise RuntimeError(f"Missing sample file for scope={scope}")

            preview_data = run_step(
                f"preview-{scope}",
                lambda scope=scope: _preview_import(client, settings.api_base, auth_token, scope, files[scope]),
            )
            expected_scope = expected.get(scope) or {}
            if int(preview_data.get("invalid_rows") or 0) != int(expected_scope.get("invalid") or 0):
                raise RuntimeError(
                    f"{scope} preview invalid_rows mismatch: got={preview_data.get('invalid_rows')}, "
                    f"expected={expected_scope.get('invalid')}"
                )

            import_data = run_step(
                f"import-{scope}",
                lambda scope=scope: _import_data(client, settings.api_base, auth_token, scope, files[scope]),
            )
            _assert_import_result(scope, import_data, expected_scope)

        summary = run_step(
            "summary-after-import",
            lambda: checked_json(
                client.get(
                    f"{settings.api_base}/xbk/analysis/summary",
                    params={"year": year, "term": term},
                    headers=auth_headers(auth_token),
                )
            ),
        )
        if int(summary.get("students") or 0) < 42:
            raise RuntimeError(f"Expected students >= 42, got {summary.get('students')}")
        if int(summary.get("courses") or 0) < 14:
            raise RuntimeError(f"Expected courses >= 14, got {summary.get('courses')}")
        selections_count = int(summary.get("selections") or 0)
        unselected_count = int(summary.get("unselected_count") or 0)
        students_count = int(summary.get("students") or 0)
        if selections_count <= 0:
            raise RuntimeError(f"Expected selections > 0, got {summary.get('selections')}")
        # Guard against accidental count amplification caused by bad join logic.
        if students_count > 0 and selections_count > students_count * 5:
            raise RuntimeError(
                f"Summary selections looks inflated: selections={selections_count}, students={students_count}"
            )
        if students_count > 0 and unselected_count > students_count * 5:
            raise RuntimeError(
                f"Summary unselected_count looks inflated: unselected_count={unselected_count}, students={students_count}"
            )

        course_results = run_step(
            "list-course-results",
            lambda: checked_json(
                client.get(
                    f"{settings.api_base}/xbk/data/course-results",
                    params={"year": year, "term": term, "page": 1, "size": 50},
                    headers=auth_headers(auth_token),
                )
            ),
        )
        if int(course_results.get("total") or 0) <= 0:
            raise RuntimeError("Expected course-results total > 0")
        items = course_results.get("items")
        if not isinstance(items, list) or not items:
            raise RuntimeError("Expected course-results items is a non-empty list")

        entities = manifest.get("entities") or {}

        update_student_no = str(entities.get("update_student_no") or "")
        students = run_step(
            "list-students-for-update",
            lambda: _list_items(
                client,
                settings.api_base,
                auth_token,
                "/xbk/data/students",
                year=year,
                term=term,
                search_text=update_student_no,
            ),
        )
        student_row = _find_exact(students, "student_no", update_student_no)
        student_payload = {
            "year": int(student_row["year"]),
            "term": str(student_row["term"]),
            "grade": student_row.get("grade"),
            "class_name": str(student_row["class_name"]),
            "student_no": str(student_row["student_no"]),
            "name": f"{student_row['name']}-PUT",
            "gender": student_row.get("gender"),
        }
        updated_student = run_step(
            "update-student",
            lambda: checked_json(
                client.put(
                    f"{settings.api_base}/xbk/data/students/{student_row['id']}",
                    json=student_payload,
                    headers=auth_headers(auth_token),
                )
            ),
        )
        if not str(updated_student.get("name") or "").endswith("-PUT"):
            raise RuntimeError("Student update assertion failed.")

        update_course_code = str(entities.get("update_course_code") or "")
        courses = run_step(
            "list-courses-for-update",
            lambda: _list_items(
                client,
                settings.api_base,
                auth_token,
                "/xbk/data/courses",
                year=year,
                term=term,
                search_text=update_course_code,
            ),
        )
        course_row = _find_exact(courses, "course_code", update_course_code)
        course_payload = {
            "year": int(course_row["year"]),
            "term": str(course_row["term"]),
            "grade": course_row.get("grade"),
            "course_code": str(course_row["course_code"]),
            "course_name": f"{course_row['course_name']}-PUT",
            "teacher": course_row.get("teacher"),
            "quota": int(course_row.get("quota") or 0) + 1,
            "location": course_row.get("location"),
        }
        updated_course = run_step(
            "update-course",
            lambda: checked_json(
                client.put(
                    f"{settings.api_base}/xbk/data/courses/{course_row['id']}",
                    json=course_payload,
                    headers=auth_headers(auth_token),
                )
            ),
        )
        if not str(updated_course.get("course_name") or "").endswith("-PUT"):
            raise RuntimeError("Course update assertion failed.")

        update_selection = entities.get("update_selection") or {}
        selection_student_no = str(update_selection.get("student_no") or "")
        selection_course_code = str(update_selection.get("course_code") or "")
        selection_rows = run_step(
            "list-selections-for-update",
            lambda: _list_items(
                client,
                settings.api_base,
                auth_token,
                "/xbk/data/selections",
                year=year,
                term=term,
                search_text=selection_student_no,
            ),
        )
        selection_row = None
        for row in selection_rows:
            if int(row.get("id") or 0) > 0 and str(row.get("course_code") or "") == selection_course_code:
                selection_row = row
                break
        if not selection_row:
            raise RuntimeError("Cannot find selection row for update.")

        selection_payload = {
            "year": int(selection_row["year"]),
            "term": str(selection_row["term"]),
            "grade": selection_row.get("grade"),
            "student_no": str(selection_row["student_no"]),
            "name": f"{selection_row.get('name') or ''}-PUT",
            "course_code": str(selection_row["course_code"]),
        }
        updated_selection = run_step(
            "update-selection",
            lambda: checked_json(
                client.put(
                    f"{settings.api_base}/xbk/data/selections/{selection_row['id']}",
                    json=selection_payload,
                    headers=auth_headers(auth_token),
                )
            ),
        )
        if "-PUT" not in str(updated_selection.get("name") or ""):
            raise RuntimeError("Selection update assertion failed.")

        new_student_nos = entities.get("new_student_nos") or []
        new_course_codes = entities.get("new_course_codes") or []
        if not new_student_nos or not new_course_codes:
            raise RuntimeError("Manifest missing new entities for delete check.")

        delete_student_no = str(new_student_nos[0])
        delete_course_code = str(new_course_codes[0])
        target_rows = _list_items(
            client,
            settings.api_base,
            auth_token,
            "/xbk/data/selections",
            year=year,
            term=term,
            search_text=delete_student_no,
        )
        target_selection = None
        for row in target_rows:
            if int(row.get("id") or 0) > 0 and str(row.get("course_code") or "") == delete_course_code:
                target_selection = row
                break
        if not target_selection:
            raise RuntimeError("Cannot find selection row for delete check.")

        run_step(
            "delete-selection",
            lambda: _delete_selection(client, settings.api_base, auth_token, int(target_selection["id"])),
        )

        check_rows = _list_items(
            client,
            settings.api_base,
            auth_token,
            "/xbk/data/selections",
            year=year,
            term=term,
            search_text=delete_student_no,
        )
        deleted_still_exists = any(
            int(row.get("id") or 0) > 0 and str(row.get("course_code") or "") == delete_course_code
            for row in check_rows
        )
        if deleted_still_exists:
            raise RuntimeError("Selection delete assertion failed: row still exists.")

        ensure_dir(settings.export_dir)
        export_scopes = {
            "students": ["年份", "学号", "姓名"],
            "courses": ["年份", "课程代码", "课程名称"],
            "selections": ["年份", "学号", "课程代码"],
            "course_results": ["年份", "学号", "课程代码", "课程名称"],
            "unselected": ["年份", "学号", "姓名"],
            "suspended": ["年份", "学号", "姓名"],
        }

        export_result: Dict[str, Any] = {}
        for scope, required_headers in export_scopes.items():
            resp = client.get(
                f"{settings.api_base}/xbk/export",
                params={"scope": scope, "year": year, "term": term, "format": "xlsx"},
                headers=auth_headers(auth_token),
            )
            resp.raise_for_status()
            file_path = settings.export_dir / f"{scope}.xlsx"
            file_path.write_bytes(resp.content)
            headers, row_count = _validate_xlsx(resp.content, required_headers)
            export_result[scope] = {
                "file": str(file_path),
                "row_count": row_count,
                "headers": headers,
            }

        run_step("exports-validate", lambda: export_result)

    report["status"] = "passed"
    return report


def main() -> None:
    parser = argparse.ArgumentParser(description="Run XBK smoke checks: import/export/update/delete.")
    parser.add_argument(
        "--manifest",
        type=str,
        default="",
        help="Path to import manifest generated by scripts/xbk/import_samples.py",
    )
    args = parser.parse_args()

    settings = get_settings()
    report_path = settings.test_result_dir / "xbk-smoke-report.json"
    compatibility_report_path = settings.test_result_dir.parent / "xbk-smoke-report.json"

    try:
        report = run_smoke(Path(args.manifest) if args.manifest else None)
        dump_json(report_path, report)
        dump_json(compatibility_report_path, report)
        print("[xbk-smoke] passed")
        print(json.dumps(report, ensure_ascii=False, indent=2))
    except Exception as exc:  # noqa: BLE001
        failed = {
            "timestamp": utc_now_iso(),
            "status": "failed",
            "error": str(exc),
        }
        dump_json(report_path, failed)
        dump_json(compatibility_report_path, failed)
        print("[xbk-smoke] failed")
        print(json.dumps(failed, ensure_ascii=False, indent=2))
        raise SystemExit(1)


if __name__ == "__main__":
    main()
